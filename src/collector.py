# src/collector.py
from __future__ import annotations
import os, sys, time, errno, msvcrt, requests, openpyxl
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

load_dotenv()
API_TOKEN = os.getenv("SKETCHFAB_TOKEN")
if not API_TOKEN:
    raise RuntimeError("SKETCHFAB_TOKEN missing (set it in .env).")
HEADERS = {"Authorization": f"Token {API_TOKEN}"}

ALIASES = {"girl": "Female", "legend of zelda": "Zelda", "links awakening": "Zelda"}
SINGLE_ASSIGNMENT_COLLECTIONS = {"hands", "gauntlets", "feet", "shoes"}

DATA_DIR = "data"
os.makedirs(DATA_DIR, exist_ok=True)
XL_PATH = os.path.join(DATA_DIR, "sketchfab_data.xlsx")

def _check_file_not_open(filepath: str) -> None:
    try:
        with open(filepath, "r+b") as f:
            msvcrt.locking(f.fileno(), msvcrt.LK_NBLCK, 1)
            msvcrt.locking(f.fileno(), msvcrt.LK_UNLCK, 1)
    except OSError as e:
        if e.errno == errno.EACCES:
            print(f"✖ '{filepath}' is open. Close it and rerun."); sys.exit(1)

def _auto_format(ws) -> None:
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 60)
    ws.freeze_panes = "A2"

def _remove_trailing_empty(ws) -> None:
    while ws.max_row > 1 and all(cell.value in (None, "") for cell in ws[ws.max_row]):
        ws.delete_rows(ws.max_row)
    while ws.max_column > 1 and all(ws.cell(row=r, column=ws.max_column).value in (None, "") for r in range(1, ws.max_row + 1)):
        ws.delete_cols(ws.max_column)

def _load_existing_columns(path=XL_PATH) -> dict:
    if not os.path.exists(path): return {}
    wb = openpyxl.load_workbook(path)
    if "Liked Models" not in wb.sheetnames: return {}
    ws = wb["Liked Models"]
    headers = [c.value for c in ws[1]]
    def idx(name): return headers.index(name) if name in headers else -1
    i_assigned = idx("Assigned Collection")
    i_auto     = idx("Auto-Assigned Collection(s)")
    i_sug      = idx("Suggested Collection(s)")
    i_fuzzy    = idx("Fuzzy Matched Collection(s)")
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        uid = row[1]
        if not uid: continue
        out[uid] = {
            "assigned": row[i_assigned] if i_assigned >= 0 else "",
            "auto":     row[i_auto]     if i_auto     >= 0 else "",
            "suggested":row[i_sug]      if i_sug      >= 0 else "",
            "fuzzy":    row[i_fuzzy]    if i_fuzzy    >= 0 else "",
        }
    return out

def _load_assigned_collections(path=XL_PATH) -> dict[str, list[str]]:
    assigned = {}
    if not os.path.exists(path): return assigned
    wb = openpyxl.load_workbook(path)
    if "Liked Models" not in wb.sheetnames: return assigned
    for row in wb["Liked Models"].iter_rows(min_row=2, values_only=True):
        uid, assigned_str = row[1], row[2]
        if uid and assigned_str:
            assigned[uid] = [c.strip() for c in str(assigned_str).split(",") if c.strip()]
    return assigned

def _get(url: str):
    r = requests.get(url, headers={"Authorization": f"Token {API_TOKEN}"}, timeout=15)
    r.raise_for_status(); return r

def get_likes() -> list[dict]:
    likes, url, page = [], "https://api.sketchfab.com/v3/me/likes?per_page=100", 0
    while url:
        r = _get(url); data = r.json()
        likes.extend(data.get("results", [])); page += 1
        print(f"Fetched likes page {page}: +{len(data.get('results', []))} → total {len(likes)}")
        url = data.get("next"); time.sleep(0.1)
    return likes

def get_collections() -> list[dict]:
    cols, url, page = [], "https://api.sketchfab.com/v3/me/collections?per_page=100", 0
    while url:
        r = _get(url); data = r.json()
        cols.extend(data.get('results', [])); page += 1
        print(f"Fetched collections page {page}: +{len(data.get('results', []))} → total {len(cols)}")
        url = data.get('next'); time.sleep(0.05)
    return cols

def get_models_in_collection(uid: str) -> list[dict]:
    models, url = [], f"https://api.sketchfab.com/v3/collections/{uid}/models?per_page=100"
    while url:
        r = _get(url); data = r.json()
        models.extend(data.get('results', []))
        url = data.get('next')
    return models

def build_uid_to_collections_map() -> dict[str, list[str]]:
    print("📥 Fetching all collections and their models...")
    mapping: dict[str, list[str]] = {}
    for col in get_collections():
        for m in get_models_in_collection(col["uid"]):
            mapping.setdefault(m["uid"], []).append(col["name"])
    print("✅ Collection mapping complete."); return mapping

def _apply_aliases(text: str) -> str:
    for a, canon in ALIASES.items(): text = text.replace(a, canon)
    return text

def _fuzzy_match_collections(tags: list[dict], collections: list[str]) -> str:
    import difflib
    tag_names = [t["name"].lower() for t in tags]
    matched: set[str] = set()
    for tag in tag_names:
        for m in difflib.get_close_matches(tag, [c.lower() for c in collections], cutoff=0.7):
            for c in collections:
                if c.lower() == m: matched.add(c)
    return ", ".join(sorted(matched))

def _suggest_collections(name: str, tags: list[dict], collections: list[str]) -> str:
    name = _apply_aliases((name or "").lower())
    tag_names = [_apply_aliases(t["name"].lower()) for t in tags]
    hits = [c for c in collections if c.lower() in name or any(c.lower() in t for t in tag_names)]
    return ", ".join(hits)

def _auto_assign(tags_str: str, suggested_str: str, fuzzy_str: str) -> str:
    tags      = [t.strip().lower() for t in (tags_str or "").split(",") if t.strip()]
    suggested = [s.strip().lower() for s in (suggested_str or "").split(",") if s.strip()]
    fuzzy     = [f.strip().lower() for f in (fuzzy_str or "").split(",") if f.strip()]
    all_cands = set(tags + suggested + fuzzy)
    votes = {c: sum([c in tags, c in suggested, c in fuzzy]) for c in all_cands}
    strong = [c for c, n in votes.items() if n == 3]
    singles = [c for c in strong if c in SINGLE_ASSIGNMENT_COLLECTIONS]
    if len(singles) == 1: return singles[0]
    if len(singles) > 1:  return ""
    return ", ".join(sorted(strong))

def _get_collection_names() -> list[str]:
    return [c["name"] for c in get_collections()]

def _write_liked_models_sheet(wb: Workbook, likes, uid2cols, assigned_map, col_names) -> None:
    ws = wb.active; ws.title = "Liked Models"
    headers = [
        "Name","UID","Assigned Collection","Already In Collection(s)","Auto-Assigned Collection(s)",
        "Suggested Collection(s)","Fuzzy Matched Collection(s)","Tags","Author","License","Downloadable"
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        if "Collection" in (cell.value or ""):
            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    prev_cols = _load_existing_columns()
    for m in likes:
        name = m.get("name"); uid = m.get("uid"); tags = m.get("tags", [])
        tag_str = ", ".join(t["name"] for t in tags)
        already_in = ", ".join(uid2cols.get(uid, []))
        suggested = _suggest_collections(name, tags, col_names)
        fuzzy = _fuzzy_match_collections(tags, col_names)
        auto = _auto_assign(tag_str, suggested, fuzzy)

        prev = prev_cols.get(uid, {})
        assigned_val  = prev.get("assigned", "")
        suggested_val = prev.get("suggested", suggested)
        fuzzy_val     = prev.get("fuzzy", fuzzy)
        auto_val      = prev.get("auto") or auto

        lic = m.get("license") or {}
        lic_str = f"{lic.get('label','')} ({lic.get('slug','')})".strip().strip("()").replace("()", "")
        is_dl = m.get("isDownloadable", False)
        author = (m.get("user") or {}).get("displayName", "")
        url = m.get("viewerUrl") or f"https://sketchfab.com/3d-models/{uid}"

        ws.append([name, uid, assigned_val, already_in, auto_val, suggested_val, fuzzy_val,
                   tag_str, author, lic_str, "Yes" if is_dl else "No"])
        ws.cell(row=ws.max_row, column=1).hyperlink = url

    _auto_format(ws)

def _write_collections_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Collections")
    ws.append(["Collection Name","Collection UID","Model Count","Model Names"])
    for col in get_collections():
        models = get_models_in_collection(col["uid"])
        names = sorted([m.get("name","") for m in models])
        ws.append([col["name"], col["uid"], len(names), ", ".join(names)])
        ws.cell(row=ws.max_row, column=1).hyperlink = f"https://sketchfab.com/collections/{col['uid']}"
    _auto_format(ws)

def build_workbook() -> str:
    _check_file_not_open(XL_PATH)
    likes = get_likes()
    uid2cols = build_uid_to_collections_map()
    assigned = _load_assigned_collections()
    col_names = _get_collection_names()

    wb = Workbook()
    _write_liked_models_sheet(wb, likes, uid2cols, assigned, col_names)
    _write_collections_sheet(wb)
    for sheet in wb.worksheets:
        _remove_trailing_empty(sheet)

    wb.save(XL_PATH)
    print(f"✔ Saved to {XL_PATH}")
    return XL_PATH

if __name__ == "__main__":
    build_workbook()
