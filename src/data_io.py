from __future__ import annotations
import os
import logging
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

DATA_DIR = os.environ.get("DATA_DIR", "data")
XL_PATH = os.path.join(DATA_DIR, "sketchfab_data.xlsx")

LIKED_SHEET = "Liked Models"
COLL_SHEET = "Collections"

os.makedirs(DATA_DIR, exist_ok=True)


def write_workbook(liked_df: pd.DataFrame, collections_df: pd.DataFrame) -> None:
    """Write workbook without dropping custom columns.
    - Preserves any extra columns present in input DataFrames.
    - Applies light formatting (Tags list -> comma string) if applicable.
    - Auto-sizes columns and trims trailing blanks.
    - If input follows the pipeline schema (has "Model UID"), tries to bring forward
      select user-edited columns from an existing workbook.
    """
    liked_out = liked_df.copy()

    # Normalize Tags column if present
    if "Tags" in liked_out.columns:
        try:
            liked_out["Tags"] = liked_out["Tags"].apply(lambda x: ", ".join(x) if isinstance(x, list) else (x or ""))
        except Exception:
            pass

    cols_out = collections_df.copy()

    # Optional: bring forward pipeline-specific columns when using that schema
    if os.path.exists(XL_PATH) and ("Model UID" in liked_out.columns):
        try:
            existing = pd.ExcelFile(XL_PATH)
            if LIKED_SHEET in existing.sheet_names:
                prev = existing.parse(LIKED_SHEET)
                if ("Model UID" in prev.columns):
                    carry_cols = [
                        "Suggested Collection(s)",
                        "Fuzzy Match Collection(s)",
                        "Assigned Collection(s)",
                        "Assignment Notes",
                    ]
                    prev_idx = prev.set_index("Model UID")
                    for col in carry_cols:
                        if col in prev_idx.columns and col in liked_out.columns:
                            m = prev_idx[col].to_dict()
                            liked_out[col] = liked_out["Model UID"].map(m)
        except Exception as e:
            logger.warning("Could not preserve previous columns: %s", e)

    with pd.ExcelWriter(XL_PATH, engine="openpyxl") as xw:
        liked_out.to_excel(xw, index=False, sheet_name=LIKED_SHEET)
        cols_out.to_excel(xw, index=False, sheet_name=COLL_SHEET)

    _finalize_workbook(XL_PATH, [LIKED_SHEET, COLL_SHEET])


def _finalize_workbook(path: str, sheets: list[str]) -> None:
    wb = load_workbook(path)
    for name in sheets:
        if name not in wb.sheetnames:
            continue
        ws = wb[name]
        # Trim trailing blank rows
        max_row = ws.max_row
        max_col = ws.max_column
        while max_row > 1 and all((ws.cell(row=max_row, column=c).value in (None, "")) for c in range(1, max_col + 1)):
            ws.delete_rows(max_row)
            max_row -= 1
        # Trim trailing blank cols
        while max_col > 1 and all((ws.cell(row=r, column=max_col).value in (None, "")) for r in range(1, max_row + 1)):
            ws.delete_cols(max_col)
            max_col -= 1
        # Auto-size columns (cap width at 60)
        for c in range(1, ws.max_column + 1):
            letter = get_column_letter(c)
            max_len = 0
            for r in range(1, ws.max_row + 1):
                v = ws.cell(row=r, column=c).value
                ln = len(str(v)) if v is not None else 0
                if ln > max_len:
                    max_len = ln
            ws.column_dimensions[letter].width = min(max_len + 2, 60)
    wb.save(path)


def read_workbook() -> tuple[pd.DataFrame, pd.DataFrame]:
    if not os.path.exists(XL_PATH):
        raise FileNotFoundError(f"Workbook not found: {XL_PATH}")
    x = pd.ExcelFile(XL_PATH)
    liked = x.parse(LIKED_SHEET)
    cols = x.parse(COLL_SHEET)
    return liked, cols
